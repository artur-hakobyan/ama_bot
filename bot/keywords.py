"""Keyword inventory: read the pillar-topic workbook, pick the next target keyword.

The workbook is the manual research output (Google Ads Keyword Planner export).
Numbers arrive in both German ("1.234,5", "Hoch") and English ("1234.5", "High")
formats depending on which tab they were pasted into, so every parse tolerates both.
"""
import re
from dataclasses import dataclass

from openpyxl import load_workbook

MASTER_TAB = "All Keywords"

# Competition labels, German and English, mapped to a sortable rank.
COMPETITION_RANK = {
    "niedrig": 0, "low": 0,
    "mittel": 1, "medium": 1,
    "hoch": 2, "high": 2,
    "unbekannt": 3, "unknown": 3, "": 3,
}


@dataclass(frozen=True)
class Keyword:
    keyword: str
    volume: int
    competition: str
    competition_index: int
    pillar: str

    @property
    def label(self) -> str:
        vol = f"{self.volume:,}".replace(",", ".") if self.volume else "?"
        return f"{self.keyword} — {vol}/Monat, {self.competition or 'unbekannt'}"


def parse_number(value) -> int:
    """Accept 5000, 5000.0, '5.000', '5000,0', '' -> int (0 when unknown)."""
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text or text == "--":
        return 0
    # Both conventions appear in this workbook:
    #   German  "5.000" / "1.234,5"  -> "." groups thousands, "," is decimal
    #   English "5000.0"             -> "." is decimal
    if "," in text and "." in text:          # "1.234,5" -> 1234.5
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:                        # "5000,0" -> 5000.0
        text = text.replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(\.\d{3})+", text):   # "5.000" -> 5000 (never 5.0)
        text = text.replace(".", "")
    try:
        return int(float(text))
    except ValueError:
        return 0


def competition_rank(label: str) -> int:
    return COMPETITION_RANK.get((label or "").strip().lower(), 3)


class KeywordSheet:
    def __init__(self, path: str):
        self._wb = load_workbook(path, data_only=True)

    @property
    def pillars(self) -> list:
        """Pillar tabs, excluding the master tab, in workbook (priority) order."""
        return [n for n in self._wb.sheetnames if n != MASTER_TAB]

    def keywords(self, pillar: str) -> list:
        ws = self._wb[pillar]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []
        cols = {str(h).strip(): i for i, h in enumerate(header) if h}
        i_kw = cols.get("Keyword", 0)
        i_vol = cols.get("Avg. monthly searches")
        i_comp = cols.get("Competition")
        i_idx = cols.get("Competition (indexed value)")
        i_pillar = cols.get("Pillar-Topics")

        out = []
        for row in rows:
            if not row or not row[i_kw]:
                continue
            out.append(Keyword(
                keyword=str(row[i_kw]).strip(),
                volume=parse_number(row[i_vol]) if i_vol is not None else 0,
                competition=str(row[i_comp] or "").strip() if i_comp is not None else "",
                competition_index=parse_number(row[i_idx]) if i_idx is not None else 0,
                pillar=(str(row[i_pillar]).strip()
                        if i_pillar is not None and row[i_pillar] else pillar),
            ))
        return out

    def ranked(self, pillar: str, used: set) -> list:
        """Highest volume first, then lowest competition; used keywords removed."""
        fresh = [k for k in self.keywords(pillar) if k.keyword.lower() not in used]
        return sorted(fresh, key=lambda k: (-k.volume, competition_rank(k.competition),
                                            k.competition_index))

    def next_keyword(self, pillar: str, used: set):
        ranked = self.ranked(pillar, used)
        return ranked[0] if ranked else None

    def remaining(self, pillar: str, used: set) -> int:
        return len(self.ranked(pillar, used))


class LiveKeywordSheet(KeywordSheet):
    """Keyword inventory read from the live Google Sheet.

    Same interface as the workbook version, so callers are unaffected. Values are
    cached briefly: a Monday run touches every tab, and the sheet rarely changes
    mid-run.
    """

    def __init__(self, google_client, spreadsheet_id: str, cache_seconds: int = 300):
        self._g = google_client
        self._id = spreadsheet_id
        self._cache_seconds = cache_seconds
        self._tabs = None
        self._rows = {}
        self._fetched = {}

    @property
    def pillars(self) -> list:
        if self._tabs is None:
            self._tabs = self._g.sheet_tabs(self._id)
        return [t for t in self._tabs if t != MASTER_TAB]

    def _values(self, pillar: str) -> list:
        import time
        now = time.time()
        if (pillar not in self._rows
                or now - self._fetched.get(pillar, 0) > self._cache_seconds):
            self._rows[pillar] = self._g.sheet_values(self._id, pillar)
            self._fetched[pillar] = now
        return self._rows[pillar]

    def keywords(self, pillar: str) -> list:
        rows = self._values(pillar)
        if not rows:
            return []
        header = rows[0]
        cols = {str(h).strip(): i for i, h in enumerate(header) if h}
        i_kw = cols.get("Keyword", 0)
        i_vol = cols.get("Avg. monthly searches")
        i_comp = cols.get("Competition")
        i_idx = cols.get("Competition (indexed value)")
        i_pillar = cols.get("Pillar-Topics")

        def cell(row, index):
            if index is None or index >= len(row):
                return ""
            return row[index]

        out = []
        for row in rows[1:]:
            if not row or not cell(row, i_kw):
                continue
            out.append(Keyword(
                keyword=str(cell(row, i_kw)).strip(),
                volume=parse_number(cell(row, i_vol)),
                competition=str(cell(row, i_comp) or "").strip(),
                competition_index=parse_number(cell(row, i_idx)),
                pillar=str(cell(row, i_pillar) or pillar).strip() or pillar,
            ))
        return out
